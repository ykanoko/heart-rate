import pandas as pd
from openpyxl import load_workbook
import os
import json
import re

# Change
data_date = "20231110_2053_"
now = "1110"


data_type = ['済', '表示方法_済']

txt_dir = 'C:\\Users\\kanok\\Documents\\大学院\\研究室\\修論\\予備実験230719~\\予備実験本番230927~\\心拍解析\\心拍切り取りデータ\\'
input_dir = "C:\\Users\\kanok\\Documents\\大学院\\研究室\\修論\\予備実験230719~\\予備実験本番230927~\\心拍解析\\作業場\\"
output_dir = 'C:\\Users\\kanok\\Documents\\大学院\\研究室\\修論\\予備実験230719~\\予備実験本番230927~\\心拍解析\\解析結果\\'

for type in data_type:
    with open(txt_dir + data_date + type + '.txt', 'r') as f:
        file_names = json.load(f)

    for file_name in file_names:
        # CSVファイルの読み込み
        input_file_name = file_name + '.SegTable'
        input_file = input_dir + input_file_name

        df_csv = pd.read_csv(input_file, encoding='shift-jis', nrows=12)
        # print("df_csv",df_csv)

        # Excelファイルの読み込み
        pattern = r'S(\d+)_'
        subject = re.search(pattern, file_name).group(1)
        sheet_index = int(file_name.split('.')[0][-1]) - 1

        output_file_name = '2309S'+ subject +'.xlsx'
        output_file = output_dir + now + '\\' + output_file_name
        wb = load_workbook(output_file)

        sheet_name = wb.sheetnames[sheet_index] # シート名を取得
        sheet = wb[sheet_name] # シートを取得

        def write_list_2d(sheet, l_2d, start_row, start_col):
            for y, row in enumerate(l_2d):
                for x, cell in enumerate(row):
                    sheet.cell(row=start_row + y,
                            column=start_col + x,
                            value=l_2d[y][x])

        write_list_2d(sheet, df_csv.values.tolist(), 2, 1)

        wb.save(output_file)

        print('input_file : ', input_file_name, 'output_file : ', output_file_name, 'sheet_name : ', sheet_name)
