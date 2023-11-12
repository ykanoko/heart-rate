
from openpyxl import load_workbook
import os

sheet_name_in = 'table'
sheet_name_out = 'Sheet1'


data_date = "1110"
dir = 'C:\\Users\\kanok\\Documents\\大学院\\研究室\\修論\\予備実験230719~\\予備実験本番230927~\\心拍解析\\解析結果\\'
dir = dir + data_date + '\\'

# Excelファイルへの書き込み
output_file_name = 'HR2309.xlsx'
output_file = dir + output_file_name


for i in range(1, 19):
    subject = i
    row = i + 3

    # Excelファイルの読み込み
    input_file_name = '2309S'+ str(subject) +'.xlsx'
    input_file = dir + input_file_name

    # print('subject : ', i, 'row', row, 'input_file : ', input_file)
    wb_in = load_workbook(filename=input_file, read_only=True, data_only=True)
    sheet_in = wb_in[sheet_name_in]
    # print('sheet_in', list(sheet_in.values))

    data = [cell.value for cell in sheet_in['B4':'EO4'][0]]

    # Excelファイルへの書き込み
    wb_out = load_workbook(output_file)
    sheet_out = wb_out[sheet_name_out]

    for j, value in enumerate(data, start=2):
        sheet_out.cell(row, column=j, value=value)

    wb_out.save(output_file)
    # print('output_file : ', output_file)



